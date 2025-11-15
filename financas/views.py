from django.shortcuts import render, redirect
from django.views import View
from django.views.generic import CreateView
from django.contrib.auth.views import LoginView, LogoutView, PasswordResetView, PasswordResetConfirmView
from django.contrib.auth.models import User
from django.urls import reverse_lazy
from django.contrib import messages
from django.db import IntegrityError
from django.db.models import Sum, Q
from decimal import Decimal
from .models import Profile
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.auth import login
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from .models import Movimentacao, Categoria, TipoMovimentacao
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime, timedelta
from django.contrib.auth.decorators import login_required 
from .forms import (
    CustomLoginForm,
    CustomUserCreationForm,
    CustomPasswordResetForm,
    CustomSetPasswordForm,
    ProfileForm  
)

import io
import json
from datetime import datetime
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl import load_workbook
import os
from django.conf import settings

from reportlab.platypus import Image
from openpyxl.styles import Border, Side
import openpyxl
import base64

# --- Login ---
class CustomLoginView(LoginView):
    template_name = "financas/login.html"
    authentication_form = CustomLoginForm

# --- Logout ---
class CustomLogoutView(LogoutView):
    next_page = reverse_lazy("landing")

def RegisterView(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            # Criar usuário usando o email como username
            email = form.cleaned_data['email']
            
            try:
                # Verificação extra para garantir que o email não existe
                if User.objects.filter(username=email).exists():
                    form.add_error('email', 'Este e-mail já está cadastrado.')
                elif User.objects.filter(email=email).exists():
                    form.add_error('email', 'Este e-mail já está cadastrado.')
                else:
                    user = User.objects.create_user(
                        username=email,  # Usa o email como username
                        email=email,
                        password=form.cleaned_data['password1'],
                        first_name=form.cleaned_data['first_name'],
                        last_name=form.cleaned_data['last_name']
                    )
                    
                    # Criar perfil com telefone
                    Profile.objects.create(
                        user=user,
                        phone=form.cleaned_data['phone']
                    )
                    
                    # Redireciona para a página de confirmação
                    return render(request, 'financas/register_done.html')
                    
            except IntegrityError as e:
                form.add_error('email', 'Erro ao criar conta. Este e-mail já pode estar em uso.')
                print(f"Erro de integridade: {e}")
                
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'financas/register.html', {'form': form})


def register_done(request):
    return render(request, 'financas/register_done.html')

# --- Landing page ---
def landing(request):
    return render(request, "financas/landing.html")

# --- Recuperação de senha ---
class CustomPasswordResetView(PasswordResetView):
    template_name = "financas/password_reset.html"
    email_template_name = "financas/password_reset_email.html"
    success_url = reverse_lazy("password_reset_done")
    form_class = CustomPasswordResetForm

# --- Redefinir senha (nova senha) ---
class CustomPasswordResetConfirmView(PasswordResetConfirmView):
    template_name = "financas/password_reset_confirm.html"
    success_url = reverse_lazy("login")
    form_class = CustomSetPasswordForm
  
# --- Tela de Transaçoes ---
@login_required
def transacoes(request):
    """View para a página de Transações"""
    # Busca todas as transações do usuário ordenadas por data
    todas_transacoes = Movimentacao.objects.filter(
        usuario=request.user
    ).select_related('categoria', 'categoria__tipoMovimentacao').order_by('-data_movimentacao')
    
    # Agrupa transações por data
    transacoes_por_data = {}
    for transacao in todas_transacoes:
        data_str = transacao.data_movimentacao.strftime('%Y-%m-%d')
        if data_str not in transacoes_por_data:
            transacoes_por_data[data_str] = []
        transacoes_por_data[data_str].append(transacao)
    
    context = {
        'transacoes_por_data': transacoes_por_data,
        'todas_transacoes': todas_transacoes
    }
    return render(request, 'financas/transacoes.html', context)

@login_required
def minha_carteira_view(request):
    periodo = request.GET.get('periodo', 'este_mes')
    
    # Buscar transações do período
    data_inicio, data_fim = calcular_periodo(periodo)
    
    transacoes_periodo = Movimentacao.objects.filter(
        usuario=request.user,
        data_movimentacao__gte=data_inicio,
        data_movimentacao__lte=data_fim
    ).select_related('categoria', 'categoria__tipoMovimentacao').order_by('-data_movimentacao')[:10]  # Limita para performance
    
    # Últimas transações (últimos 30 dias)
    data_limite_transacoes = timezone.now().date() - timedelta(days=30)
    ultimas_transacoes = Movimentacao.objects.filter(
        usuario=request.user,
        data_movimentacao__gte=data_limite_transacoes
    ).select_related('categoria', 'categoria__tipoMovimentacao').order_by('-data_movimentacao')[:5]
    
    # Calcular totais
    saldo_info = calcular_saldo_usuario(request.user, periodo)
    investimentos_info = calcular_investimentos_usuario(request.user, periodo)
    
    # Buscar categorias para os modais
    categorias_receita = Categoria.categorias_por_tipo('RECEITA')
    categorias_despesa = Categoria.categorias_por_tipo('DESPESA') 
    categorias_economia = Categoria.categorias_por_tipo('ECONOMIA')
    
    # Datas para templates
    hoje = timezone.now().date()
    ontem = hoje - timedelta(days=1)
    
    context = {
        'transacoes_periodo': transacoes_periodo,
        'ultimas_transacoes': ultimas_transacoes,
        'saldo': saldo_info['saldo'],
        'saldo_variacao': saldo_info['variacao'],
        'saldo_porcentagem': saldo_info['porcentagem'],
        'investimentos': investimentos_info['valor'],
        'invest_variacao': investimentos_info['variacao'],
        'invest_porcentagem': investimentos_info['porcentagem'],
        'periodo_selecionado': periodo,
        'data_hoje': hoje.strftime('%Y-%m-%d'),
        'data_ontem': ontem.strftime('%Y-%m-%d'),
        'periodo_texto': obter_texto_periodo(periodo),
        'categorias_receita': categorias_receita,
        'categorias_despesa': categorias_despesa,
        'categorias_economia': categorias_economia,
    }
    return render(request, 'financas/carteira.html', context)

def obter_texto_periodo(periodo):
    textos = {
        'este_mes': 'Este mês',
        'ultimo_mes': 'Último mês', 
        'ultimos_3_meses': 'Últimos 3 meses'
    }
    return textos.get(periodo, 'este mês')


# Funções auxiliares
def calcular_periodo(periodo):
    """Calcula as datas de início e fim baseado no período selecionado"""
    hoje = timezone.now().date()
    
    if periodo == 'este_mes':
        data_inicio = hoje.replace(day=1)
        data_fim = hoje
    elif periodo == 'ultimo_mes':
        primeiro_dia_mes_atual = hoje.replace(day=1)
        data_fim = primeiro_dia_mes_atual - timedelta(days=1)
        data_inicio = data_fim.replace(day=1)
    elif periodo == 'ultimos_3_meses':
        data_fim = hoje
        data_inicio = hoje - timedelta(days=90)
    else:  # default: este mês
        data_inicio = hoje.replace(day=1)
        data_fim = hoje
    
    return data_inicio, data_fim

def calcular_saldo_usuario(usuario, periodo):
    """Calcula saldo, variação e porcentagem para o período do usuário"""
    data_inicio, data_fim = calcular_periodo(periodo)
    
    # Buscar período anterior para comparação
    if periodo == 'este_mes':
        periodo_anterior = 'ultimo_mes'
    elif periodo == 'ultimo_mes':
        periodo_anterior = 'ultimos_3_meses'
    else:  # últimos 3 meses
        periodo_anterior = 'ultimo_mes'
    
    data_inicio_anterior, data_fim_anterior = calcular_periodo(periodo_anterior)
    
    # Calcular saldo atual
    receitas_atual = Movimentacao.objects.filter(
        usuario=usuario,
        data_movimentacao__gte=data_inicio,
        data_movimentacao__lte=data_fim,
        categoria__tipoMovimentacao__nome='RECEITA'
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0')
    
    despesas_atual = Movimentacao.objects.filter(
        usuario=usuario,
        data_movimentacao__gte=data_inicio,
        data_movimentacao__lte=data_fim,
        categoria__tipoMovimentacao__nome='DESPESA'
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0')
    
    saldo_atual = receitas_atual - despesas_atual
    
    # Calcular saldo período anterior
    receitas_anterior = Movimentacao.objects.filter(
        usuario=usuario,
        data_movimentacao__gte=data_inicio_anterior,
        data_movimentacao__lte=data_fim_anterior,
        categoria__tipoMovimentacao__nome='RECEITA'
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0')
    
    despesas_anterior = Movimentacao.objects.filter(
        usuario=usuario,
        data_movimentacao__gte=data_inicio_anterior,
        data_movimentacao__lte=data_fim_anterior,
        categoria__tipoMovimentacao__nome='DESPESA'
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0')
    
    saldo_anterior = receitas_anterior - despesas_anterior
    
    # Calcular variação
    if saldo_anterior != 0:
        variacao_valor = saldo_atual - saldo_anterior
        variacao_porcentagem = (variacao_valor / abs(saldo_anterior)) * 100
    else:
        variacao_valor = saldo_atual
        variacao_porcentagem = 100 if saldo_atual > 0 else -100 if saldo_atual < 0 else 0
    
    return {
        'saldo': saldo_atual,
        'variacao': variacao_valor,
        'porcentagem': variacao_porcentagem
    }

def calcular_investimentos_usuario(usuario, periodo):
    """Calcula investimentos, variação e porcentagem para o período do usuário"""
    data_inicio, data_fim = calcular_periodo(periodo)
    
    # Buscar período anterior para comparação
    if periodo == 'este_mes':
        periodo_anterior = 'ultimo_mes'
    elif periodo == 'ultimo_mes':
        periodo_anterior = 'ultimos_3_meses'
    else:
        periodo_anterior = 'ultimo_mes'
    
    data_inicio_anterior, data_fim_anterior = calcular_periodo(periodo_anterior)
    
    # Calcular investimentos atual
    investimentos_atual = Movimentacao.objects.filter(
        usuario=usuario,
        data_movimentacao__gte=data_inicio,
        data_movimentacao__lte=data_fim,
        categoria__tipoMovimentacao__nome='ECONOMIA'
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0')
    
    # Calcular investimentos período anterior
    investimentos_anterior = Movimentacao.objects.filter(
        usuario=usuario,
        data_movimentacao__gte=data_inicio_anterior,
        data_movimentacao__lte=data_fim_anterior,
        categoria__tipoMovimentacao__nome='ECONOMIA'
    ).aggregate(total=Sum('valor'))['total'] or Decimal('0')
    
    # Calcular variação
    if investimentos_anterior != 0:
        variacao_valor = investimentos_atual - investimentos_anterior
        variacao_porcentagem = (variacao_valor / abs(investimentos_anterior)) * 100
    else:
        variacao_valor = investimentos_atual
        variacao_porcentagem = 100 if investimentos_atual > 0 else -100 if investimentos_atual < 0 else 0
    
    return {
        'valor': investimentos_atual,
        'variacao': variacao_valor,
        'porcentagem': variacao_porcentagem
    }

@login_required
def dashboard(request):
    return render(request, "financas/dashboard.html")

# @login_required  
# def relatorios(request):
#     return render(request, "financas/relatorios.html")

@login_required  
def relatorios(request):
    from datetime import date
    
    # Datas padrão: primeiro dia do mês atual até hoje
    data_final = date.today()
    data_inicial = data_final.replace(day=1)  # Primeiro dia do mês
    
    # Buscar categorias para os dropdowns
    categorias_receita = Categoria.categorias_por_tipo('RECEITA')
    categorias_despesa = Categoria.categorias_por_tipo('DESPESA') 
    categorias_economia = Categoria.categorias_por_tipo('ECONOMIA')
    
    context = {
        'categorias_receita': categorias_receita,
        'categorias_despesa': categorias_despesa,
        'categorias_economia': categorias_economia,
        'data_inicial': data_inicial.strftime('%Y-%m-%d'),
        'data_final': data_final.strftime('%Y-%m-%d'),
    }
    return render(request, "financas/relatorios.html", context)


@login_required
def minha_conta(request):
    profile, created = Profile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = ProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            # Atualiza o perfil
            form.save()
            
            # Atualiza os dados do usuário
            user = request.user
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.email = form.cleaned_data['email']
            user.save()
            
            messages.success(request, 'Perfil atualizado com sucesso!')
            return redirect('minha-conta')
    else:
        # Preenche o formulário com os dados atuais
        form = ProfileForm(initial={
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'email': request.user.email,
            'phone': profile.phone,
        })
    
    return render(request, 'financas/minha_conta.html', {
        'form': form,
        'profile': profile
    })


#Views de finanças

@login_required
def get_categorias_por_tipo(request, tipo_nome):
    """Retorna categorias filtradas por tipo"""
    categorias = Categoria.categorias_por_tipo(tipo_nome)
    categorias_list = [
        {'id': cat.id_categoria, 'nome': cat.nomeCategoria}
        for cat in categorias
    ]
    return JsonResponse({'categorias': categorias_list})

@login_required
def criar_movimentacao(request):
    if request.method == 'POST':
        try:
            categoria_id = request.POST.get('categoria_id')
            data_movimentacao = request.POST.get('data')
            valor = request.POST.get('valor')
            descricao = request.POST.get('descricao')
            
            movimentacao = Movimentacao(
                usuario=request.user,
                data_movimentacao=data_movimentacao,
                valor=valor,
                descricao=descricao,
                categoria_id=categoria_id
            )
            movimentacao.save()
            
            messages.success(request, 'Movimentação criada com sucesso!')
            return redirect('minha-carteira')
            
        except Exception as e:
            messages.error(request, f'Erro ao criar movimentação: {str(e)}')
            return redirect('minha-carteira')
    
    return redirect('minha-carteira')






@method_decorator(csrf_exempt, name='dispatch')
class ExportarRelatorioView(View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            # Obter parâmetros do frontend
            data_inicial = data.get('data_inicial')
            data_final = data.get('data_final')
            categorias_ganhos = data.get('categorias_ganhos', [])
            categorias_despesas = data.get('categorias_despesas', [])
            categorias_investimentos = data.get('categorias_investimentos', [])
            formato = data.get('formato', 'pdf')
            
            # Filtrar transações baseado nos parâmetros
            transacoes = Movimentacao.objects.filter(usuario=request.user)
            
            if data_inicial and data_final:
                transacoes = transacoes.filter(
                    data_movimentacao__range=[data_inicial, data_final]
                )
            
            # Filtrar por categorias selecionadas
            categorias_selecionadas = []
            if categorias_ganhos:
                categorias_selecionadas.extend(categorias_ganhos)
            if categorias_despesas:
                categorias_selecionadas.extend(categorias_despesas)
            if categorias_investimentos:
                categorias_selecionadas.extend(categorias_investimentos)
            
            if categorias_selecionadas:
                transacoes = transacoes.filter(categoria_id__in=categorias_selecionadas)
            
            # Ordenar por data (mais recente primeiro)
            transacoes = transacoes.order_by('-data_movimentacao')
            
            if formato == 'pdf':
                return self.gerar_pdf(transacoes, data_inicial, data_final, request.user)
            elif formato == 'xlsx':
                return self.gerar_excel(transacoes, data_inicial, data_final, request.user)
            else:
                return HttpResponse('Formato não suportado', status=400)
                
        except Exception as e:
            return HttpResponse(f'Erro ao gerar relatório: {str(e)}', status=500)
    
    # def gerar_pdf(self, transacoes, data_inicial, data_final, usuario):
    #     buffer = io.BytesIO()
    #     doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=30)
    #     elements = []
    #     styles = getSampleStyleSheet()
        
    #     # Estilo personalizado para o título
    #     title_style = ParagraphStyle(
    #         'CustomTitle',
    #         parent=styles['Heading1'],
    #         fontSize=16,
    #         spaceAfter=30,
    #         alignment=1,  # Centralizado
    #     )
        
    #     # Título
    #     title = Paragraph("RELATÓRIO DE TRANSAÇÕES - KINGCOIN", title_style)
    #     elements.append(title)
        
    #     # Informações do período
    #     # periodo_text = f"Período: {data_inicial} a {data_final}" if data_inicial and data_final else "Período: Todos"
    #     def formatar_data_brasileira(data_str):
    #         if data_str:
    #             try:
    #                 data_obj = datetime.strptime(data_str, '%Y-%m-%d')
    #                 return data_obj.strftime('%d/%m/%Y')
    #             except:
    #                  return data_str
    #         return data_str

    #     data_inicial_br = formatar_data_brasileira(data_inicial)
    #     data_final_br = formatar_data_brasileira(data_final)

    #     periodo_text = f"Período: {data_inicial_br} a {data_final_br}" if data_inicial and data_final else "Período: Todos"
        
    #     periodo_style = ParagraphStyle(
    #         'PeriodoStyle',
    #         parent=styles['Normal'],
    #         fontSize=10,
    #         spaceAfter=20,
    #     )
    #     periodo = Paragraph(periodo_text, periodo_style)
    #     elements.append(periodo)
        
    #     # Tabela de transações
    #     if transacoes.exists():
    #         # Cabeçalho da tabela
    #         data = [['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor (R$)']]
            
    #         for transacao in transacoes:
    #             tipo = transacao.categoria.tipoMovimentacao.nome
    #             valor = f"R$ {transacao.valor:.2f}"
                
    #             # Adiciona sinal baseado no tipo
    #             if tipo in ['DESPESA', 'ECONOMIA']:
    #                 valor = f"-{valor}"
    #             else:
    #                 valor = f"+{valor}"
                
    #             data.append([
    #                 transacao.data_movimentacao.strftime('%d/%m/%Y'),
    #                 transacao.descricao,
    #                 transacao.categoria.nomeCategoria,
    #                 tipo,
    #                 valor
    #             ])
            
    #         # Criar tabela
    #         table = Table(data)
    #         table.setStyle(TableStyle([
    #             ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#23b785')),
    #             ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    #             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
    #             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    #             ('FONTSIZE', (0, 0), (-1, 0), 10),
    #             ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    #             ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    #             ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
    #             ('FONTSIZE', (0, 1), (-1, -1), 8),
    #             ('GRID', (0, 0), (-1, -1), 1, colors.black)
    #         ]))
            
    #         elements.append(table)
            
    #         # Resumo
    #         elements.append(Spacer(1, 20))
    #         total_receitas = sum(t.valor for t in transacoes if t.categoria.tipoMovimentacao.nome == 'RECEITA')
    #         total_despesas = sum(t.valor for t in transacoes if t.categoria.tipoMovimentacao.nome == 'DESPESA')
    #         total_investimentos = sum(t.valor for t in transacoes if t.categoria.tipoMovimentacao.nome == 'ECONOMIA')
    #         saldo = total_receitas - total_despesas - total_investimentos
            
    #         resumo_data = [
    #             ['RESUMO', ''],
    #             ['Total de Receitas', f'R$ {total_receitas:.2f}'],
    #             ['Total de Despesas', f'R$ {total_despesas:.2f}'],
    #             ['Total Investido', f'R$ {total_investimentos:.2f}'],
    #             ['Saldo Disponível', f'R$ {saldo:.2f}']
    #         ]
            
    #         resumo_table = Table(resumo_data, colWidths=[200, 100])
    #         resumo_table.setStyle(TableStyle([
    #             ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#23b785')),
    #             ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
    #             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    #             ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
    #             ('GRID', (0, 0), (-1, -1), 1, colors.black)
    #         ]))
            
    #         elements.append(resumo_table)
            
    #     else:
    #         elements.append(Paragraph("Nenhuma transação encontrada para os filtros selecionados.", styles['Normal']))
        
    #     # Rodapé
    #     elements.append(Spacer(1, 30))
    #     rodape = Paragraph(f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} para {usuario.username}", styles['Italic'])
    #     elements.append(rodape)
        
    #     # Gerar PDF
    #     doc.build(elements)
    #     buffer.seek(0)
        
    #     response = HttpResponse(buffer, content_type='application/pdf')
    #     response['Content-Disposition'] = f'attachment; filename="relatorio_kingcoin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    #     return response
    
    # def gerar_excel(self, transacoes, data_inicial, data_final, usuario):
    #     wb = Workbook()
    #     ws = wb.active
    #     ws.title = "Relatório Transações"
        
    #     # Estilos
    #     header_font = Font(bold=True, color="FFFFFF")
    #     header_fill = PatternFill(start_color="23b785", end_color="23b785", fill_type="solid")
    #     center_align = Alignment(horizontal='center')
        
    #     # Cabeçalho
    #     ws.merge_cells('A1:E1')
    #     ws['A1'] = "RELATÓRIO DE TRANSAÇÕES - KINGCOIN"
    #     ws['A1'].font = Font(bold=True, size=14)
    #     ws['A1'].alignment = center_align
        
    #     # ws['A2'] = f"Período: {data_inicial} a {data_final}" if data_inicial and data_final else "Período: Todos"
    #     def formatar_data_brasileira(data_str):
    #         if data_str:
    #             try:
    #                 data_obj = datetime.strptime(data_str, '%Y-%m-%d')
    #                 return data_obj.strftime('%d/%m/%Y')
    #             except:
    #                 return data_str
    #         return data_str

    #     data_inicial_br = formatar_data_brasileira(data_inicial)
    #     data_final_br = formatar_data_brasileira(data_final)

    #     ws['A2'] = f"Período: {data_inicial_br} a {data_final_br}" if data_inicial and data_final else "Período: Todos"     



    #     ws['A2'].font = Font(italic=True)
        
    #     # Cabeçalho da tabela
    #     headers = ['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor (R$)']
    #     for col, header in enumerate(headers, 1):
    #         cell = ws.cell(row=4, column=col, value=header)
    #         cell.font = header_font
    #         cell.fill = header_fill
    #         cell.alignment = center_align
        
    #     # Dados das transações
    #     row = 5
    #     for transacao in transacoes:
    #         tipo = transacao.categoria.tipoMovimentacao.nome
    #         valor = transacao.valor
            
    #         ws.cell(row=row, column=1, value=transacao.data_movimentacao.strftime('%d/%m/%Y'))
    #         ws.cell(row=row, column=2, value=transacao.descricao)
    #         ws.cell(row=row, column=3, value=transacao.categoria.nomeCategoria)
    #         ws.cell(row=row, column=4, value=tipo)
            
    #         cell_valor = ws.cell(row=row, column=5, value=float(valor))
    #         cell_valor.number_format = '"R$ "#,##0.00'
            
    #         # Colorir baseado no tipo
    #         if tipo in ['DESPESA', 'ECONOMIA']:
    #             cell_valor.font = Font(color="FF0000")  # Vermelho para saídas
    #         else:
    #             cell_valor.font = Font(color="008000")  # Verde para entradas
            
    #         row += 1
        
    #     # Resumo
    #     if transacoes.exists():
    #         row += 2
    #         ws.cell(row=row, column=1, value="RESUMO").font = Font(bold=True)
            
    #         total_receitas = sum(float(t.valor) for t in transacoes if t.categoria.tipoMovimentacao.nome == 'RECEITA')
    #         total_despesas = sum(float(t.valor) for t in transacoes if t.categoria.tipoMovimentacao.nome == 'DESPESA')
    #         total_investimentos = sum(float(t.valor) for t in transacoes if t.categoria.tipoMovimentacao.nome == 'ECONOMIA')
    #         saldo = total_receitas - total_despesas - total_investimentos
            
    #         resumo_data = [
    #             ("Total de Receitas", total_receitas),
    #             ("Total de Despesas", total_despesas),
    #             ("Total Investido", total_investimentos),
    #             ("Saldo Disponível", saldo)
    #         ]
            
    #         for i, (label, valor) in enumerate(resumo_data, 1):
    #             ws.cell(row=row+i, column=1, value=label)
    #             cell = ws.cell(row=row+i, column=2, value=float(valor))
    #             cell.number_format = '"R$ "#,##0.00'
        
    #     # Ajustar largura das colunas
    #     for column in ['A', 'B', 'C', 'D', 'E']:
    #         ws.column_dimensions[column].width = 15
        
    #     ws.column_dimensions['B'].width = 30  # Descrição mais larga
        
    #     # Rodapé
    #     row += len(resumo_data) + 3 if transacoes.exists() else row + 2
    #     ws.cell(row=row, column=1, value=f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} para {usuario.username}").font = Font(italic=True)
        
    #     # Salvar para response
    #     buffer = io.BytesIO()
    #     wb.save(buffer)
    #     buffer.seek(0)
        
    #     response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    #     response['Content-Disposition'] = f'attachment; filename="relatorio_kingcoin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    #     return response

    def gerar_pdf(self, transacoes, data_inicial, data_final, usuario):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=80)
        elements = []
        styles = getSampleStyleSheet()
        
        # --- CABEÇALHO COM IDENTIDADE VISUAL ---
        # Logo em texto estilizado (mais confiável que SVG)
        logo_style = ParagraphStyle(
            'LogoStyle',
            parent=styles['Heading1'],
            fontSize=28,
            textColor=colors.HexColor('#23b785'),
            alignment=1,  # Centralizado
            spaceAfter=5,
            fontName='Helvetica-Bold',
            textTransform='uppercase'
        )
        elements.append(Paragraph("KINGCOIN", logo_style))
        
        # Símbolo da moeda (representação gráfica)
        symbol_style = ParagraphStyle(
            'SymbolStyle',
            parent=styles['Normal'],
            fontSize=14,
            textColor=colors.HexColor('#23b785'),
            alignment=1,
            spaceAfter=15
        )
        elements.append(Paragraph("● Sistema de Gestão Financeira ●", symbol_style))
        
        # Linha decorativa
        line_style = ParagraphStyle(
            'LineStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#23b785'),
            alignment=1,
            spaceAfter=20
        )
        elements.append(Paragraph("•" * 50, line_style))
        
        # Título do relatório
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=25,
            alignment=1,
            textColor=colors.black,
            fontName='Helvetica-Bold'
        )
        title = Paragraph("RELATÓRIO DE TRANSAÇÕES FINANCEIRAS", title_style)
        elements.append(title)
        
        # Informações do período
        def formatar_data_brasileira(data_str):
            if data_str:
                try:
                    data_obj = datetime.strptime(data_str, '%Y-%m-%d')
                    return data_obj.strftime('%d/%m/%Y')
                except:
                    return data_str
            return data_str

        data_inicial_br = formatar_data_brasileira(data_inicial)
        data_final_br = formatar_data_brasileira(data_final)

        periodo_text = f"Período: {data_inicial_br} a {data_final_br}" if data_inicial and data_final else "Período: Todos"
        
        periodo_style = ParagraphStyle(
            'PeriodoStyle',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=25,
            alignment=1,
            textColor=colors.gray,
            fontName='Helvetica'
        )
        periodo = Paragraph(periodo_text, periodo_style)
        elements.append(periodo)
        
        # Tabela de transações
        if transacoes.exists():
            # Cabeçalho da tabela
            data = [['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor (R$)']]
            
            for transacao in transacoes:
                tipo = transacao.categoria.tipoMovimentacao.nome
                valor_str = f"R$ {transacao.valor:,.2f}"
                
                # Adiciona sinal baseado no tipo
                if tipo in ['DESPESA', 'ECONOMIA']:
                    valor_str = f"- {valor_str}"
                else:
                    valor_str = f"+ {valor_str}"
                
                data.append([
                    transacao.data_movimentacao.strftime('%d/%m/%Y'),
                    transacao.descricao[:30] + '...' if len(transacao.descricao) > 30 else transacao.descricao,
                    transacao.categoria.nomeCategoria,
                    tipo,
                    valor_str
                ])
            
            # Criar tabela
            table = Table(data, colWidths=[60, 150, 80, 70, 90])
            table.setStyle(TableStyle([
                # Cabeçalho
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#23b785')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Linhas de dados
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            
            # Resumo financeiro
            elements.append(Spacer(1, 25))
            
            # Calcular totais
            total_receitas = sum(t.valor for t in transacoes if t.categoria.tipoMovimentacao.nome == 'RECEITA')
            total_despesas = sum(t.valor for t in transacoes if t.categoria.tipoMovimentacao.nome == 'DESPESA')
            total_investimentos = sum(t.valor for t in transacoes if t.categoria.tipoMovimentacao.nome == 'ECONOMIA')
            saldo = total_receitas - total_despesas - total_investimentos
            
            # Título do resumo
            resumo_title_style = ParagraphStyle(
                'ResumoTitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=12,
                alignment=0,
                textColor=colors.HexColor('#23b785'),
                fontName='Helvetica-Bold'
            )
            elements.append(Paragraph("RESUMO FINANCEIRO", resumo_title_style))
            
            resumo_data = [
                ['Total de Receitas', f'R$ {total_receitas:,.2f}'],
                ['Total de Despesas', f'R$ {total_despesas:,.2f}'],
                ['Total Investido/ Poupado', f'R$ {total_investimentos:,.2f}'],
                ['SALDO DISPONÍVEL', f'R$ {saldo:,.2f}']
            ]
            
            resumo_table = Table(resumo_data, colWidths=[200, 120])
            resumo_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e8f5e8')),
                ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#ffe8e8')),
                ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#e8f0ff')),
                ('BACKGROUND', (0, 3), (-1, 3), colors.HexColor('#23b785')),
                ('TEXTCOLOR', (0, 3), (-1, 3), colors.white),
                ('FONTNAME', (0, 0), (-1, 2), 'Helvetica'),
                ('FONTNAME', (0, 3), (-1, 3), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            
            elements.append(resumo_table)
            
        else:
            # Mensagem quando não há transações
            no_data_style = ParagraphStyle(
                'NoDataStyle',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.gray,
                alignment=1,
                spaceAfter=20
            )
            elements.append(Paragraph("Nenhuma transação encontrada para os filtros selecionados.", no_data_style))
        
        # Rodapé
        elements.append(Spacer(1, 30))
        rodape_style = ParagraphStyle(
            'RodapeStyle',
            parent=styles['Italic'],
            fontSize=8,
            alignment=1,
            textColor=colors.gray
        )
        rodape_text = f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} • Usuário: {usuario.username} • KingCoin Sistema Financeiro"
        rodape = Paragraph(rodape_text, rodape_style)
        elements.append(rodape)
        
        # Gerar PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="relatorio_kingcoin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    

    def gerar_excel(self, transacoes, data_inicial, data_final, usuario):
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório KingCoin"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="23b785", end_color="23b785", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        green_font = Font(color="008000", bold=True)
        red_font = Font(color="FF0000", bold=True)
        blue_font = Font(color="0000FF", bold=True)
        kingcoin_font = Font(bold=True, size=20, color="23b785")
        title_font = Font(bold=True, size=14)
        resumo_header_font = Font(bold=True, color="23b785", size=11)
        
        # --- CABEÇALHO COM IDENTIDADE VISUAL ---
        # Linha 1: Logo KingCoin
        ws.merge_cells('A1:E1')
        ws['A1'] = "KINGCOIN"
        ws['A1'].font = kingcoin_font
        ws['A1'].alignment = center_align
        
        # Linha 2: Símbolo gráfico
        ws.merge_cells('A2:E2')
        ws['A2'] = "● Sistema de Gestão Financeira ●"
        ws['A2'].font = Font(color="23b785", size=11)
        ws['A2'].alignment = center_align
        
        # Linha 3: Título do relatório
        ws.merge_cells('A3:E3')
        ws['A3'] = "RELATÓRIO DE TRANSAÇÕES FINANCEIRAS"
        ws['A3'].font = title_font
        ws['A3'].alignment = center_align
        
        # Linha 4: Período
        def formatar_data_brasileira(data_str):
            if data_str:
                try:
                    data_obj = datetime.strptime(data_str, '%Y-%m-%d')
                    return data_obj.strftime('%d/%m/%Y')
                except:
                    return data_str
            return data_str

        data_inicial_br = formatar_data_brasileira(data_inicial)
        data_final_br = formatar_data_brasileira(data_final)

        ws.merge_cells('A4:E4')
        ws['A4'] = f"Período: {data_inicial_br} a {data_final_br}" if data_inicial and data_final else "Período: Todos"
        ws['A4'].font = Font(italic=True, size=11)
        ws['A4'].alignment = center_align
        
        # Linha 5: Espaço em branco
        ws.merge_cells('A5:E5')
        ws['A5'] = ""
        
        # Cabeçalho da tabela (linha 6)
        headers = ['Data', 'Descrição', 'Categoria', 'Tipo', 'Valor (R$)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Dados das transações
        row = 7
        for transacao in transacoes:
            tipo = transacao.categoria.tipoMovimentacao.nome
            valor = float(transacao.valor)
            
            # Data
            ws.cell(row=row, column=1, value=transacao.data_movimentacao.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=1).alignment = center_align
            
            # Descrição
            ws.cell(row=row, column=2, value=transacao.descricao)
            ws.cell(row=row, column=2).alignment = left_align
            
            # Categoria
            ws.cell(row=row, column=3, value=transacao.categoria.nomeCategoria)
            ws.cell(row=row, column=3).alignment = center_align
            
            # Tipo
            ws.cell(row=row, column=4, value=tipo)
            ws.cell(row=row, column=4).alignment = center_align
            
            # Valor
            cell_valor = ws.cell(row=row, column=5, value=valor)
            cell_valor.number_format = '"R$ "#,##0.00'
            cell_valor.alignment = right_align
            
            # Colorir baseado no tipo
            if tipo in ['DESPESA']:
                cell_valor.font = red_font
            elif tipo == 'ECONOMIA':
                cell_valor.font = blue_font
            else:
                cell_valor.font = green_font

            # Adicionar bordas às células
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Alternar cores de fundo para as linhas (zebrado)
            if row % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F8F9FA", 
                        end_color="F8F9FA", 
                        fill_type="solid"
                    )
            
            row += 1
        
        # Resumo financeiro
        if transacoes.exists():
            row += 2
            
            # Título do resumo
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = "RESUMO FINANCEIRO"
            ws[f'A{row}'].font = resumo_header_font
            ws[f'A{row}'].alignment = left_align
            row += 1
            
            # Calcular totais
            total_receitas = sum(float(t.valor) for t in transacoes if t.categoria.tipoMovimentacao.nome == 'RECEITA')
            total_despesas = sum(float(t.valor) for t in transacoes if t.categoria.tipoMovimentacao.nome == 'DESPESA')
            total_investimentos = sum(float(t.valor) for t in transacoes if t.categoria.tipoMovimentacao.nome == 'ECONOMIA')
            saldo = total_receitas - total_despesas - total_investimentos
            
            resumo_data = [
                ("Total de Receitas", total_receitas, green_font),
                ("Total de Despesas", total_despesas, red_font),
                ("Total Investido/ Poupado", total_investimentos, Font(color="0000FF", bold=True)),
                ("SALDO DISPONÍVEL", saldo, Font(bold=True, color="23b785" if saldo >= 0 else "FF0000", size=12))
            ]
            
            for i, (label, valor, font_style) in enumerate(resumo_data):
                # Label
                ws.merge_cells(f'A{row+i}:B{row+i}')
                ws.cell(row=row+i, column=1, value=label)
                ws.cell(row=row+i, column=1).font = Font(bold=True)
                ws.cell(row=row+i, column=1).alignment = left_align
                
                # Valor
                ws.merge_cells(f'C{row+i}:E{row+i}')
                cell = ws.cell(row=row+i, column=3, value=valor)
                cell.number_format = '"R$ "#,##0.00'
                cell.font = font_style
                cell.alignment = right_align
                
                # Adicionar borda a todas as células da linha
                for col in [1, 2, 3, 4, 5]:
                    ws.cell(row=row+i, column=col).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                # Cor de fundo para a linha do saldo
                if label == "SALDO FINAL":
                    for col in [1, 2, 3, 4, 5]:
                        ws.cell(row=row+i, column=col).fill = PatternFill(
                            start_color="E8F5E8" if saldo >= 0 else "FFE8E8", 
                            end_color="E8F5E8" if saldo >= 0 else "FFE8E8", 
                            fill_type="solid"
                        )
            
            row += len(resumo_data)
        
        else:
            # Mensagem quando não há transações
            row += 2
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = "Nenhuma transação encontrada para os filtros selecionados."
            ws[f'A{row}'].font = Font(italic=True, color="666666")
            ws[f'A{row}'].alignment = center_align
            row += 1
        
        # Rodapé
        row += 2
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')} • Usuário: {usuario.username} • KingCoin Sistema Financeiro"
        ws[f'A{row}'].font = Font(italic=True, size=8, color="666666")
        ws[f'A{row}'].alignment = center_align
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 12   # Data
        ws.column_dimensions['B'].width = 35   # Descrição
        ws.column_dimensions['C'].width = 20   # Categoria
        ws.column_dimensions['D'].width = 12   # Tipo
        ws.column_dimensions['E'].width = 15   # Valor
        
        # Ajustar altura das linhas do cabeçalho
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[6].height = 25
        
        # Congelar painel (cabeçalho fixo)
        ws.freeze_panes = 'A7'
        
        # Salvar para response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="relatorio_kingcoin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response